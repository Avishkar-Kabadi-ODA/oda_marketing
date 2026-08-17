# Copyright (c) 2026, Optimum Data Analytics and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import add_days, getdate, nowdate, now_datetime, get_url


class ContentItem(Document):
	def validate(self):
		self.validate_description_length()
		self.validate_assigned_to_reviewer_mutual_exclusion()
		self.validate_creation_permissions()
		self.validate_metadata_edit_permissions()
		self.validate_writer_readonly_before_briefed()
		self.validate_reviewer_readonly_fields()
		self.validate_workflow_transition_permissions()
		self.sync_status_with_workflow()
		self.validate_publish_date_with_calendar()
		self.check_overdue_sla()
		self.validate_primary_attachment_mandatory()
		self.validate_technical_reviewer_mandatory()
		self.validate_revision_notes()
		self.validate_published_url_mandatory()
		self.stamp_brief_accepted_on()

	def before_workflow_action(self, action=None):
		if frappe.flags.ignore_permissions or self.flags.ignore_permissions or getattr(self.flags, "ignore_workflow", False):
			return

		user = frappe.session.user
		if user == "Administrator":
			return

		user_lower = (user or "").lower()
		assigned_lower = (self.assigned_to or "").lower()
		reviewer_lower = (self.reviewer_technical or "").lower()

		is_doc_writer = (user_lower == assigned_lower)
		is_doc_reviewer = (user_lower == reviewer_lower)
		roles = frappe.get_roles(user)
		is_lead = ("Marketing Lead" in roles or "System Manager" in roles) and not is_doc_writer

		if action in ["Request Changes", "Approve"]:
			if is_doc_writer:
				frappe.throw(
					_("As the assigned writer, you cannot review, approve, or request changes on your own deliverable."),
					frappe.PermissionError
				)
			if not is_doc_reviewer and not is_lead:
				frappe.throw(
					_("Only the assigned Reviewer (<b>{0}</b>) or a Marketing Lead can review this deliverable.").format(self.reviewer_technical or "Reviewer"),
					frappe.PermissionError
				)

		elif action in ["Start Work", "Submit for Review", "Resubmit Draft"]:
			if is_doc_reviewer and not is_doc_writer and not is_lead:
				frappe.throw(
					_("Only the assigned Writer (<b>{0}</b>) can perform drafting actions on this deliverable.").format(self.assigned_to or "Writer"),
					frappe.PermissionError
				)

		elif action in ["Publish"]:
			from oda_marketing.permissions import get_default_publisher
			publisher = get_default_publisher()
			is_publisher = user_lower == (publisher or "").lower()
			if not (is_lead or is_publisher):
				frappe.throw(_("Only Marketing Leads and Publishers can mark deliverables as Published."), frappe.PermissionError)

	def validate_workflow_transition_permissions(self):
		if frappe.flags.ignore_permissions or self.flags.ignore_permissions or getattr(self.flags, "ignore_workflow", False):
			return
		before = self.get_doc_before_save()
		if not before:
			return

		curr_state = getattr(self, "workflow_state", None) or getattr(self, "status", None)
		prev_state = getattr(before, "workflow_state", None) or getattr(before, "status", None)

		if not curr_state or not prev_state or curr_state == prev_state:
			return

		user = frappe.session.user
		if user == "Administrator":
			return

		user_lower = (user or "").lower()
		assigned_lower = (self.assigned_to or "").lower()
		reviewer_lower = (self.reviewer_technical or "").lower()

		is_doc_writer = (user_lower == assigned_lower)
		is_doc_reviewer = (user_lower == reviewer_lower)
		roles = frappe.get_roles(user)
		is_lead = ("Marketing Lead" in roles or "System Manager" in roles) and not is_doc_writer

		# Review & Approval transitions: ONLY reviewer or Lead (who is NOT the writer)
		if prev_state == "In Review" and curr_state in ["In Revision", "Approved"]:
			if is_doc_writer:
				frappe.throw(
					_("As the assigned writer, you cannot review, approve, or request changes on your own deliverable."),
					frappe.PermissionError
				)
			if not is_doc_reviewer and not is_lead:
				frappe.throw(
					_("Only the assigned Reviewer (<b>{0}</b>) or a Marketing Lead can review this deliverable.").format(self.reviewer_technical or "Reviewer"),
					frappe.PermissionError
				)

		# Drafting / Resubmission transitions: ONLY writer or Lead
		if prev_state in ["Briefed", "In Revision"] and curr_state == "In Progress":
			if is_doc_reviewer and not is_doc_writer and not is_lead:
				frappe.throw(
					_("Only the assigned Writer (<b>{0}</b>) can accept the brief and work on drafts.").format(self.assigned_to or "Writer"),
					frappe.PermissionError
				)

		if prev_state == "In Progress" and curr_state == "In Review":
			if is_doc_reviewer and not is_doc_writer and not is_lead:
				frappe.throw(
					_("Only the assigned Writer (<b>{0}</b>) can submit this deliverable for review.").format(self.assigned_to or "Writer"),
					frappe.PermissionError
				)

		# Publishing transition: ONLY Lead / Publisher (not writer)
		if prev_state == "Approved" and curr_state == "Published":
			from oda_marketing.permissions import get_default_publisher
			publisher = get_default_publisher()
			is_publisher = user_lower == (publisher or "").lower()
			if not (is_lead or is_publisher):
				frappe.throw(_("Only Marketing Leads and Publishers can mark deliverables as Published."), frappe.PermissionError)

	def before_insert(self):
		if self.assigned_to:
			self.owner = self.assigned_to
		self.sync_status_with_workflow()
		if not getattr(self, "reminder_1_days_before", None):
			settings = frappe.get_single("Marketing Settings")
			if getattr(settings, "sla_reminder_enabled", 0):
				self.reminder_1_days_before = int(getattr(settings, "sla_reminder_days_before", 3) or 3)

	def validate_assigned_to_reviewer_mutual_exclusion(self):
		"""Ensures the assigned writer cannot review their own content."""
		if self.assigned_to and self.reviewer_technical:
			if self.assigned_to.strip().lower() == self.reviewer_technical.strip().lower():
				frappe.throw(
					_("The <b>Assigned To</b> writer cannot be the same user as the <b>Reviewer</b>. A creator cannot review their own deliverable."),
					frappe.ValidationError
				)

	def get_user_effective_role(self, user=None):
		"""Determines the user's role specifically in the context of THIS document."""
		if not user:
			user = frappe.session.user
		if user == "Administrator":
			return "Lead"
		user_lower = (user or "").lower()
		assigned_lower = (self.assigned_to or "").lower()
		reviewer_lower = (self.reviewer_technical or "").lower()

		# Document-level authorship takes precedence: an author is strictly a Writer for this doc
		if user_lower == assigned_lower:
			return "Writer"
		if user_lower == reviewer_lower:
			return "Reviewer"

		roles = frappe.get_roles(user)
		if "Marketing Lead" in roles or "System Manager" in roles:
			return "Lead"
		if "Technical Reviewer" in roles:
			return "Reviewer"
		return "Writer"

	def validate_description_length(self):
		if self.description and len(str(self.description).strip()) > 500:
			frappe.throw(
				_("<b>Description</b> exceeds maximum limit of 500 characters. (Current length: {0} characters)").format(len(str(self.description).strip())),
				frappe.ValidationError
			)


	def is_lead_user(self):
		user = frappe.session.user
		if user == "Administrator":
			return True
		roles = frappe.get_roles(user)
		return "Marketing Lead" in roles or "System Manager" in roles

	def is_reviewer_user(self):
		user = frappe.session.user
		if user == "Administrator":
			return True
		roles = frappe.get_roles(user)
		return "Technical Reviewer" in roles

	def validate_creation_permissions(self):
		if frappe.flags.ignore_permissions or self.flags.ignore_permissions:
			return
		if self.is_new() and not self.is_lead_user():
			frappe.throw(_("Only <b>Marketing Leads</b> can create new Content Items."), frappe.PermissionError)

	def validate_metadata_edit_permissions(self):
		if frappe.flags.ignore_permissions or self.flags.ignore_permissions:
			return
		effective_role = self.get_user_effective_role()
		if not self.is_new() and effective_role != "Lead":
			before = self.get_doc_before_save()
			if not before:
				return
			# due_date is explicitly excluded — Content Writers can edit Due Date
			metadata_fields = [
				"title", "content_type", "description", "industry_domain",
				"content_calendar", "planned_publish_date", "assigned_to",
				"reviewer_technical"
			]
			for field in metadata_fields:
				val_self = getattr(self, field, None)
				val_before = getattr(before, field, None)

				if field in ["planned_publish_date"]:
					if val_self and val_before and getdate(val_self) != getdate(val_before):
						frappe.throw(_("Only <b>Marketing Leads</b> are permitted to modify core item metadata ({0}).").format(field), frappe.PermissionError)
				else:
					if str(val_self or "") != str(val_before or ""):
						frappe.throw(_("Only <b>Marketing Leads</b> are permitted to modify core item metadata ({0}).").format(field), frappe.PermissionError)

	def validate_writer_readonly_before_briefed(self):
		"""Content Writer cannot edit attachments or notes while item is in Planned state."""
		if frappe.flags.ignore_permissions or self.flags.ignore_permissions:
			return
		if self.get_user_effective_role() == "Lead":
			return
		if getattr(self, "workflow_state", None) != "Planned":
			return

		before = self.get_doc_before_save()
		if not before:
			return

		attachment_fields = ["content_file_1", "content_file_2", "content_file_3", "notes"]
		for field in attachment_fields:
			val_self = getattr(self, field, None)
			val_before = getattr(before, field, None)
			if str(val_self or "") != str(val_before or ""):
				frappe.throw(
					_("Content Writers cannot edit attachments or notes while the item is in <b>Planned</b> state. Wait until the brief is issued."),
					frappe.PermissionError
				)

	def validate_reviewer_readonly_fields(self):
		"""Technical Reviewer cannot edit attachments, notes, or metadata - only revision_feedback_notes, reviewer_copilot_instructions, and workflow_state (via workflow actions)."""
		if frappe.flags.ignore_permissions or self.flags.ignore_permissions:
			return
		effective_role = self.get_user_effective_role()
		if effective_role == "Lead":
			return
		if effective_role != "Reviewer":
			return

		before = self.get_doc_before_save()
		if not before:
			return

		# Forbidden metadata and draft attachment fields that Technical Reviewers cannot modify
		forbidden_fields = [
			"title", "content_type", "description", "industry_domain",
			"content_calendar", "planned_publish_date", "assigned_to",
			"reviewer_technical", "published_url", "risk_flag", "due_date",
			"content_file_1", "content_file_2", "content_file_3", "notes"
		]

		for fieldname in forbidden_fields:
			val_self = getattr(self, fieldname, None)
			val_before = getattr(before, fieldname, None)

			if fieldname in ["planned_publish_date", "due_date"]:
				if val_self and val_before and getdate(val_self) != getdate(val_before):
					frappe.throw(
						_("Technical Reviewers cannot edit item metadata or attachments (<b>{0}</b>). Only revision feedback, copilot instructions, and workflow actions can be modified.").format(self.meta.get_label(fieldname) or fieldname),
						frappe.PermissionError
					)
			else:
				if str(val_self or "") != str(val_before or ""):
					frappe.throw(
						_("Technical Reviewers cannot edit item metadata or attachments (<b>{0}</b>). Only revision feedback, copilot instructions, and workflow actions can be modified.").format(self.meta.get_label(fieldname) or fieldname),
						frappe.PermissionError
					)

	def sync_status_with_workflow(self):
		wf_state = getattr(self, "workflow_state", None)
		if wf_state:
			self.status = wf_state
		elif not getattr(self, "status", None):
			self.status = "Planned"

	def validate_publish_date_with_calendar(self):
		if self.content_calendar and self.planned_publish_date:
			cal = frappe.get_doc("Content Calendar", self.content_calendar)
			pub_date = getdate(self.planned_publish_date)
			if cal.from_date and pub_date < getdate(cal.from_date):
				frappe.throw(_("Planned Publish Date ({0}) cannot be before Content Calendar start date ({1}).").format(self.planned_publish_date, cal.from_date))
			if cal.to_date and pub_date > getdate(cal.to_date):
				frappe.throw(_("Planned Publish Date ({0}) cannot be after Content Calendar end date ({1}).").format(self.planned_publish_date, cal.to_date))

	def check_overdue_sla(self):
		settings = frappe.get_single("Marketing Settings")
		if getattr(settings, "enable_auto_overdue_flag", 0):
			if self.due_date and getattr(self, "workflow_state", None) not in ["Approved", "Published"]:
				if getdate(nowdate()) > getdate(self.due_date):
					self.risk_flag = "Late"

	def validate_primary_attachment_mandatory(self):
		target_states = ["In Review", "Approved", "Published"]
		if getattr(self, "workflow_state", None) in target_states and not self.content_file_1:
			frappe.throw(_("<b>Primary Content Draft (content_file_1)</b> is mandatory before submitting for review or publishing."))

	def validate_technical_reviewer_mandatory(self):
		if getattr(self, "workflow_state", None) == "In Review" and not self.reviewer_technical:
			frappe.throw(_("<b>Reviewer</b> must be assigned before submitting for Review."))

	def validate_revision_notes(self):
		if getattr(self, "workflow_state", None) == "In Revision" and not getattr(self, "revision_feedback_notes", None):
			frappe.throw(_("<b>Revision Feedback / Notes</b> are mandatory when requesting revisions."))

	def validate_published_url_mandatory(self):
		if getattr(self, "workflow_state", None) == "Published" and not getattr(self, "published_url", None):
			frappe.throw(_("<b>Published URL</b> is mandatory when marking a deliverable as Published."))

	def stamp_brief_accepted_on(self):
		"""Auto-stamp brief_accepted_on when writer transitions from Briefed to In Progress."""
		if getattr(self, "workflow_state", None) == "In Progress" and not getattr(self, "brief_accepted_on", None):
			before = self.get_doc_before_save()
			if before and getattr(before, "workflow_state", None) == "Briefed":
				self.brief_accepted_on = now_datetime()

	def get_user_full_name(self, user_email):
		if not user_email:
			return "Team Member"
		first_name, last_name = frappe.db.get_value("User", user_email, ["first_name", "last_name"]) or (None, None)
		if first_name:
			return f"{first_name} {last_name or ''}".strip()

		if "@" in user_email:
			username_part = user_email.split("@")[0]
			formatted = username_part.replace(".", " ").replace("_", " ").title()
			return formatted

		return user_email

	def get_file_link_html(self, file_path, label="View Primary Draft"):
		if not file_path:
			return ""
		full_url = get_url(file_path)
		return f'<a href="{full_url}" target="_blank" style="color: #2563eb; text-decoration: underline; font-weight: 600;">{label}</a>'

	def get_template_context(self):
		settings = frappe.get_single("Marketing Settings")
		publisher_email = settings.default_publisher if hasattr(settings, "default_publisher") else None
		item_url = get_url(f"/app/content-item/{self.name}")

		context = {
			"doc": self,
			"content_item_url": item_url,
			"creator_name": self.get_user_full_name(self.owner),
			"assigned_to_name": self.get_user_full_name(self.assigned_to),
			"reviewer_technical_name": self.get_user_full_name(self.reviewer_technical),
			"publisher_name": self.get_user_full_name(publisher_email),
			"content_file_1_link": self.get_file_link_html(self.content_file_1, "View Primary Content Draft"),
			"content_file_2_link": self.get_file_link_html(self.content_file_2, "View Supporting Asset 1"),
			"content_file_3_link": self.get_file_link_html(self.content_file_3, "View Supporting Asset 2"),
		}
		return context

	def on_update(self):
		self.trigger_workflow_notifications()
		self.trigger_system_notifications()

	def trigger_system_notifications(self):
		"""Sends targeted Frappe In-App Bell 🔔 Notifications to the specific user involved."""
		previous_doc = self.get_doc_before_save()
		previous_state = getattr(self.flags, "previous_workflow_state", None) or (previous_doc.workflow_state if previous_doc else None)
		current_state = self.workflow_state

		if previous_state == current_state or not current_state:
			return

		target_user = None
		subject = None

		if current_state == "Briefed" and self.assigned_to:
			target_user = self.assigned_to
			subject = f"Assigned Deliverable: '{self.title}' (Brief Issued)"

		elif current_state == "In Review" and self.reviewer_technical:
			target_user = self.reviewer_technical
			subject = f"Review Required: '{self.title}'"

		elif current_state == "In Revision" and self.assigned_to:
			target_user = self.assigned_to
			subject = f"Revisions Requested: '{self.title}'"

		elif current_state == "Approved":
			settings = frappe.get_single("Marketing Settings")
			publisher = getattr(settings, "default_publisher", None)
			if publisher:
				target_user = publisher
				subject = f"Approved for Publishing: '{self.title}'"

		elif current_state == "Published" and self.assigned_to:
			target_user = self.assigned_to
			subject = f"Congratulations! Deliverable Published: '{self.title}'"

		if target_user and subject:
			try:
				if frappe.db.exists("User", target_user):
					doc_n = frappe.get_doc({
						"doctype": "Notification Log",
						"subject": subject,
						"for_user": target_user,
						"type": "Alert",
						"document_type": "Content Item",
						"document_name": self.name,
						"email_content": f"Deliverable '{self.title}' status is now {current_state}."
					})
					doc_n.insert(ignore_permissions=True)
			except Exception as e:
				frappe.log_error(f"Failed to insert in-app Notification Log for {target_user}: {str(e)}")

	def trigger_workflow_notifications(self):
		settings = frappe.get_single("Marketing Settings")
		if not settings.enable_email_notifications:
			return

		previous_doc = self.get_doc_before_save()
		previous_state = getattr(self.flags, "previous_workflow_state", None) or (previous_doc.workflow_state if previous_doc else None)
		current_state = self.workflow_state

		if previous_state == current_state or not current_state:
			return

		recipients = set()
		cc = set()
		template_name = None

		publisher_email = getattr(settings, "default_publisher", None)

		if current_state == "Briefed":
			if self.assigned_to:
				recipients.add(self.assigned_to)
			if self.owner and self.owner != self.assigned_to:
				cc.add(self.owner)
			template_name = settings.writer_email_template

		elif current_state == "In Review":
			if self.reviewer_technical:
				recipients.add(self.reviewer_technical)
			if self.assigned_to and self.assigned_to != self.reviewer_technical:
				cc.add(self.assigned_to)
			if self.owner and self.owner != self.reviewer_technical:
				cc.add(self.owner)
			if publisher_email and publisher_email != self.reviewer_technical:
				cc.add(publisher_email)
			template_name = settings.reviewer_email_template

		elif current_state == "In Revision":
			if self.assigned_to:
				recipients.add(self.assigned_to)
			if self.owner and self.owner != self.assigned_to:
				cc.add(self.owner)
			template_name = settings.writer_email_template

		elif current_state == "Approved":
			if publisher_email:
				recipients.add(publisher_email)
			if self.owner and self.owner != publisher_email:
				cc.add(self.owner)
			if self.assigned_to and self.assigned_to != publisher_email:
				cc.add(self.assigned_to)
			template_name = settings.publisher_email_template

		elif current_state == "Published":
			if self.assigned_to:
				recipients.add(self.assigned_to)
			template_name = getattr(settings, "published_email_template", None)

		if recipients and template_name and frappe.db.exists("Email Template", template_name):
			try:
				tmpl = frappe.get_doc("Email Template", template_name)
				ctx = self.get_template_context()

				subject = frappe.render_template(tmpl.subject, ctx)
				message = frappe.render_template(tmpl.response, ctx)

				frappe.sendmail(
					recipients=list(recipients),
					cc=list(cc),
					subject=subject,
					message=message,
					now=False
				)
			except Exception as e:
				frappe.log_error(f"Failed to send workflow email for Content Item {self.name}: {str(e)}")


def send_overdue_sla_notifications():
	"""Scheduled job to alert involved parties for late items and upcoming due dates (Restricted to Business Hours)."""
	settings = frappe.get_single("Marketing Settings")
	if not settings.enable_email_notifications:
		return

	today = getdate(nowdate())

	# Execute email dispatch ONCE per day at Business Hours Start (default 9 AM / 09:00)
	current_hour = frappe.utils.now_datetime().hour
	start_hour = int(getattr(settings, "business_hours_start", 9) or 9)

	if current_hour != start_hour:
		return

	template_name = getattr(settings, "overdue_email_template", None) or getattr(settings, "overdue_sla_email_template", None)
	if not (template_name and frappe.db.exists("Email Template", template_name)):
		return

	tmpl = frappe.get_doc("Email Template", template_name)

	# Send alerts for overdue (Late) items (deduplicated by last_overdue_notified_on)
	overdue_items = frappe.get_all(
		"Content Item",
		filters={
			"risk_flag": "Late",
			"workflow_state": ["not in", ["Approved", "Published"]]
		},
		fields=["name", "last_overdue_notified_on"]
	)

	for item_data in overdue_items:
		# Prevent daily duplicate emails if already notified today
		if item_data.last_overdue_notified_on and getdate(item_data.last_overdue_notified_on) == today:
			continue

		item = frappe.get_doc("Content Item", item_data.name)
		recipients = set()
		if item.assigned_to:
			recipients.add(item.assigned_to)
		if item.reviewer_technical:
			recipients.add(item.reviewer_technical)

		if recipients:
			try:
				ctx = item.get_template_context()
				subject = frappe.render_template(tmpl.subject, ctx)
				message = frappe.render_template(tmpl.response, ctx)

				frappe.sendmail(
					recipients=list(recipients),
					subject=subject,
					message=message,
					now=True
				)
				item.db_set("last_overdue_notified_on", today, update_modified=False)
			except Exception as e:
				frappe.log_error(f"Failed to send overdue email for {item.name}: {str(e)}")

	# Send reminder notifications for items approaching due date (Up to 3 reminders: Marketing Settings default + 2 content-level)
	sla_reminder_enabled = getattr(settings, "sla_reminder_enabled", 0)

	active_items = frappe.get_all(
		"Content Item",
		filters={
			"risk_flag": ["!=", "Late"],
			"workflow_state": ["not in", ["Approved", "Published"]]
		},
		fields=["name", "due_date", "reminder_1_days_before", "reminder_2_days_before"]
	)

	default_days_before = int(getattr(settings, "sla_reminder_days_before", 3) or 3) if sla_reminder_enabled else None

	# Dedicated Due Date Reminder template fallback
	reminder_template_name = "Marketing Due Date Reminder" if frappe.db.exists("Email Template", "Marketing Due Date Reminder") else template_name
	reminder_tmpl = frappe.get_doc("Email Template", reminder_template_name) if reminder_template_name else tmpl

	for item_data in active_items:
		if not item_data.due_date:
			continue

		due_date = getdate(item_data.due_date)

		# Collect up to 3 reminder day offsets
		reminder_offsets = set()
		if default_days_before is not None:
			reminder_offsets.add(default_days_before)
		if item_data.reminder_1_days_before:
			reminder_offsets.add(int(item_data.reminder_1_days_before))
		if item_data.reminder_2_days_before:
			reminder_offsets.add(int(item_data.reminder_2_days_before))

		# Check if today matches any of the reminder dates
		should_remind = False
		matched_days = None
		for days in sorted(reminder_offsets):
			rem_date = add_days(due_date, -days)
			if getdate(rem_date) == today:
				should_remind = True
				matched_days = days
				break

		if should_remind:
			item = frappe.get_doc("Content Item", item_data.name)
			recipients = set()
			if item.assigned_to:
				recipients.add(item.assigned_to)
			if item.reviewer_technical:
				recipients.add(item.reviewer_technical)

			if recipients:
				try:
					ctx = item.get_template_context()
					ctx["matched_days"] = matched_days
					subject = frappe.render_template(reminder_tmpl.subject, ctx)
					message = frappe.render_template(reminder_tmpl.response, ctx)

					frappe.sendmail(
						recipients=list(recipients),
						subject=subject,
						message=message,
						now=True
					)
				except Exception as e:
					frappe.log_error(f"Failed to send reminder email for {item.name}: {str(e)}")

	frappe.db.commit()


import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@frappe.whitelist()
def get_reviewer_users(doctype=None, txt="", searchfield="name", start=0, page_len=20, filters=None):
	"""Filters Reviewer link dropdown to show active System Users holding Technical Reviewer, Marketing Lead, or System Manager role. Excludes the assigned writer if provided."""
	exclude_user = None
	if isinstance(filters, dict):
		exclude_user = filters.get("assigned_to") or filters.get("exclude_user")
	elif isinstance(filters, str):
		try:
			import json
			parsed = json.loads(filters)
			if isinstance(parsed, dict):
				exclude_user = parsed.get("assigned_to") or parsed.get("exclude_user")
		except Exception:
			pass

	conditions = [
		"u.enabled = 1",
		"u.user_type = 'System User'",
		"(r.role IN ('Technical Reviewer', 'Marketing Lead', 'System Manager') OR u.name = 'Administrator')"
	]
	params = []
	if exclude_user:
		conditions.append("u.name != %s")
		params.append(exclude_user)

	conditions.append("(u.name LIKE %s OR u.first_name LIKE %s OR u.last_name LIKE %s)")
	params.extend([f"%{txt}%", f"%{txt}%", f"%{txt}%"])
	params.extend([int(start or 0), int(page_len or 20)])

	where_clause = " AND ".join(conditions)
	return frappe.db.sql(f"""
		SELECT DISTINCT u.name, CONCAT_WS(' ', u.first_name, u.last_name)
		FROM `tabUser` u
		LEFT JOIN `tabHas Role` r ON r.parent = u.name
		WHERE {where_clause}
		ORDER BY u.name ASC
		LIMIT %s, %s
	""", tuple(params))



@frappe.whitelist()
def get_assigned_to_users(doctype=None, txt="", searchfield="name", start=0, page_len=20, filters=None):
	"""Filters Assigned To link dropdown to show active System Users. Excludes the reviewer if provided."""
	exclude_user = None
	if isinstance(filters, dict):
		exclude_user = filters.get("reviewer_technical") or filters.get("exclude_user")
	elif isinstance(filters, str):
		try:
			import json
			parsed = json.loads(filters)
			if isinstance(parsed, dict):
				exclude_user = parsed.get("reviewer_technical") or parsed.get("exclude_user")
		except Exception:
			pass

	conditions = ["u.enabled = 1", "u.user_type = 'System User'", "u.name != 'Administrator'"]
	params = []
	if exclude_user:
		conditions.append("u.name != %s")
		params.append(exclude_user)

	conditions.append("(u.name LIKE %s OR u.first_name LIKE %s OR u.last_name LIKE %s)")
	params.extend([f"%{txt}%", f"%{txt}%", f"%{txt}%"])
	params.extend([int(start or 0), int(page_len or 20)])

	where_clause = " AND ".join(conditions)
	return frappe.db.sql(f"""
		SELECT DISTINCT u.name, CONCAT_WS(' ', u.first_name, u.last_name)
		FROM `tabUser` u
		WHERE {where_clause}
		ORDER BY u.name ASC
		LIMIT %s, %s
	""", tuple(params))


@frappe.whitelist()
def download_content_item_template():
	"""Generates and downloads a standardized Excel template (.xlsx) for bulk Content Item importing."""
	if not frappe.has_permission("Content Item", "create"):
		frappe.throw(_("You do not have permission to import Content Items."), frappe.PermissionError)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Content Items"

	# Styles
	header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
	header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
	center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
	left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
	border_thin = Side(style="thin", color="CBD5E1")
	cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

	headers = [
		("Title*", 32, "Required. Title of the content deliverable."),
		("Format*", 18, "Required. Format option (e.g., Blog, Poll, Flowchart, Carousel)."),
		("Description*", 45, "Required. Topic/brief description (max 500 characters)."),
		("Industry Domain", 22, "Optional. Target domain (e.g., HCLS, Fintech, Pharma Supply Chain)."),
		("Content Calendar", 28, "Optional. Calendar name (e.g., 2026 Marketing Calendar)."),
		("Planned Publish Date*", 22, "YYYY-MM-DD. Target live publishing date."),
		("Due Date*", 20, "YYYY-MM-DD. Writer submission due date."),
		("Assigned To*", 26, "Required. Email of assigned writer/creator."),
		("Reviewer*", 26, "Required. Email of technical reviewer (Must differ from Assigned To)."),
		("Notes", 35, "Optional. Initial notes or drafting pointers.")
	]

	for col_num, (header_text, width, _) in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col_num, value=header_text)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = center_align
		cell.border = cell_border
		col_letter = get_column_letter(col_num)
		ws.column_dimensions[col_letter].width = width

	# Sample data rows
	formats = frappe.get_all("Content Item Option", filters={"option_type": "Format", "is_active": 1}, pluck="option_label")
	default_format = formats[0] if formats else "Blog"
	domains = frappe.get_all("Content Item Option", filters={"option_type": "Industry Domain", "is_active": 1}, pluck="option_label")
	default_domain = domains[0] if domains else "HCLS"
	active_cal = frappe.get_all("Content Calendar", filters={"status": "Active"}, pluck="name")
	default_cal = active_cal[0] if active_cal else ""

	sample_rows = [
		[
			"Mastering Multi-Cloud Governance in 2026",
			default_format,
			"Comprehensive guide covering enterprise multi-cloud governance and FinOps best practices.",
			default_domain,
			default_cal,
			"2026-09-15",
			"2026-09-01",
			frappe.session.user if frappe.session.user != "Administrator" else "writer@example.com",
			"reviewer@example.com",
			"Include key stats on cloud cost savings."
		],
		[
			"AI In Clinical Trials: Adoption Poll",
			"Poll" if "Poll" in formats else default_format,
			"Interactive survey on clinical trial workflow automation challenges.",
			default_domain,
			default_cal,
			"2026-09-20",
			"2026-09-10",
			"writer2@example.com",
			"reviewer@example.com",
			"Target 4 distinct choice options."
		]
	]

	sample_font = Font(name="Calibri", size=10)
	for row_idx, row_data in enumerate(sample_rows, 2):
		for col_idx, val in enumerate(row_data, 1):
			c = ws.cell(row=row_idx, column=col_idx, value=val)
			c.font = sample_font
			c.alignment = left_align
			c.border = cell_border

	ws.row_dimensions[1].height = 28
	ws.row_dimensions[2].height = 22
	ws.row_dimensions[3].height = 22

	# Reference sheet for available options
	ws_ref = wb.create_sheet(title="Allowed Options")
	ws_ref.cell(row=1, column=1, value="Active Formats").font = Font(bold=True)
	ws_ref.cell(row=1, column=2, value="Industry Domains").font = Font(bold=True)
	ws_ref.cell(row=1, column=3, value="Active Calendars").font = Font(bold=True)

	for i, f in enumerate(formats, 2):
		ws_ref.cell(row=i, column=1, value=f)
	for i, d in enumerate(domains, 2):
		ws_ref.cell(row=i, column=2, value=d)
	for i, c in enumerate(active_cal, 2):
		ws_ref.cell(row=i, column=3, value=c)

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.response["filename"] = "Content_Item_Import_Template.xlsx"
	frappe.response["filecontent"] = output.getvalue()
	frappe.response["type"] = "binary"


@frappe.whitelist()
def import_content_items_from_excel(file_url=None, default_calendar=None):
	"""Parses an uploaded .xlsx or .csv spreadsheet and imports Content Items into 'Planned' status."""
	if not frappe.has_permission("Content Item", "create"):
		frappe.throw(_("Only Marketing Leads and Administrators can import Content Items."), frappe.PermissionError)

	if not file_url:
		frappe.throw(_("Please provide an uploaded file URL to import."), frappe.ValidationError)

	# Read file from filesystem
	file_doc = frappe.get_doc("File", {"file_url": file_url}) if frappe.db.exists("File", {"file_url": file_url}) else None
	if file_doc:
		file_path = file_doc.get_full_path()
	else:
		file_path = frappe.get_site_path("public", file_url.lstrip("/"))

	import os
	if not os.path.exists(file_path):
		frappe.throw(_("File not found on server at {0}").format(file_path), frappe.DoesNotExistError)

	# Determine default calendar fallback
	if not default_calendar:
		settings = frappe.get_single("Marketing Settings")
		default_calendar = getattr(settings, "default_content_calendar", None)
		if not default_calendar or not frappe.db.exists("Content Calendar", default_calendar):
			active_cals = frappe.get_all("Content Calendar", filters={"status": "Active"}, limit=1, pluck="name")
			default_calendar = active_cals[0] if active_cals else None

	rows_data = []

	if file_path.endswith(".csv"):
		import csv
		with open(file_path, "r", encoding="utf-8-sig") as f:
			reader = csv.reader(f)
			raw_rows = list(reader)
			if raw_rows:
				header = [str(h).strip().lower().replace("*", "").replace(" ", "_") for h in raw_rows[0]]
				for r in raw_rows[1:]:
					if any(str(c).strip() for c in r):
						rows_data.append(dict(zip(header, [str(c).strip() for c in r])))
	else:
		wb = openpyxl.load_workbook(file_path, data_only=True)
		ws = wb.active
		raw_rows = list(ws.iter_rows(values_only=True))
		if raw_rows:
			raw_header = raw_rows[0]
			header = [str(h).strip().lower().replace("*", "").replace(" ", "_") if h is not None else "" for h in raw_header]
			for r in raw_rows[1:]:
				if any(c is not None and str(c).strip() for c in r):
					row_dict = {}
					for idx, col_name in enumerate(header):
						if col_name and idx < len(r):
							val = r[idx]
							if isinstance(val, (frappe.utils.datetime.date, frappe.utils.datetime.datetime)):
								val = str(val)[:10]
							row_dict[col_name] = str(val).strip() if val is not None else ""
					rows_data.append(row_dict)

	if not rows_data:
		frappe.throw(_("The uploaded file does not contain any valid data rows."), frappe.ValidationError)

	created_items = []
	errors = []

	# Cache active options
	valid_formats = set(frappe.get_all("Content Item Option", filters={"option_type": "Format", "is_active": 1}, pluck="option_label"))
	valid_domains = set(frappe.get_all("Content Item Option", filters={"option_type": "Industry Domain", "is_active": 1}, pluck="option_label"))

	for idx, row in enumerate(rows_data, start=2):
		row_num = idx
		title = row.get("title") or row.get("deliverable_title") or row.get("name")
		content_type = row.get("format") or row.get("content_type") or row.get("type")
		description = row.get("description") or row.get("topic") or row.get("brief")
		industry_domain = row.get("industry_domain") or row.get("practice_area") or row.get("domain")
		calendar = row.get("content_calendar") or row.get("calendar") or default_calendar
		planned_pub_date = row.get("planned_publish_date") or row.get("publish_date")
		due_date = row.get("due_date") or row.get("sla_due_date")
		assigned_to = row.get("assigned_to") or row.get("writer") or row.get("author")
		reviewer = row.get("reviewer") or row.get("reviewer_technical") or row.get("technical_reviewer")
		notes = row.get("notes") or row.get("remarks")

		# Validation
		row_errors = []
		if not title:
			row_errors.append("Title is mandatory")
		if not content_type:
			row_errors.append("Format is mandatory")
		elif content_type not in valid_formats:
			# Case-insensitive match check
			matched = next((f for f in valid_formats if f.lower() == content_type.lower()), None)
			if matched:
				content_type = matched
			else:
				row_errors.append(f"Invalid Format '{content_type}'. Allowed: {', '.join(sorted(valid_formats))}")

		if not description:
			row_errors.append("Description is mandatory")
		elif len(description) > 500:
			row_errors.append(f"Description exceeds 500 characters ({len(description)} chars)")

		if industry_domain and industry_domain not in valid_domains:
			matched_d = next((d for d in valid_domains if d.lower() == industry_domain.lower()), None)
			if matched_d:
				industry_domain = matched_d

		if not calendar:
			row_errors.append("Content Calendar is required (none specified or found active)")
		elif not frappe.db.exists("Content Calendar", calendar):
			row_errors.append(f"Content Calendar '{calendar}' does not exist")

		if not planned_pub_date:
			row_errors.append("Planned Publish Date is mandatory")

		if not due_date:
			if planned_pub_date:
				try:
					due_date = str(add_days(getdate(planned_pub_date), -7))
				except Exception:
					row_errors.append("Due Date is mandatory")
			else:
				row_errors.append("Due Date is mandatory")

		if not assigned_to:
			row_errors.append("Assigned To writer is mandatory")
		elif not frappe.db.exists("User", assigned_to):
			row_errors.append(f"Assigned To user '{assigned_to}' does not exist")

		if not reviewer:
			row_errors.append("Reviewer is mandatory")
		elif not frappe.db.exists("User", reviewer):
			row_errors.append(f"Reviewer user '{reviewer}' does not exist")

		if assigned_to and reviewer and assigned_to.strip().lower() == reviewer.strip().lower():
			row_errors.append(f"Assigned To and Reviewer cannot be the same user ({assigned_to})")

		if row_errors:
			errors.append({
				"row": row_num,
				"title": title or f"Row {row_num}",
				"errors": row_errors
			})
			continue

		try:
			doc = frappe.get_doc({
				"doctype": "Content Item",
				"title": title,
				"content_type": content_type,
				"description": description,
				"industry_domain": industry_domain or None,
				"content_calendar": calendar,
				"planned_publish_date": planned_pub_date,
				"due_date": due_date,
				"assigned_to": assigned_to,
				"reviewer_technical": reviewer,
				"notes": notes or None,
				"status": "Planned",
				"workflow_state": "Planned",
				"risk_flag": "On track",
				"ai_review_status": "Not Started"
			})
			doc.insert(ignore_permissions=True)
			created_items.append({
				"name": doc.name,
				"title": doc.title,
				"assigned_to": doc.assigned_to,
				"status": "Planned"
			})
		except Exception as e:
			errors.append({
				"row": row_num,
				"title": title or f"Row {row_num}",
				"errors": [str(e)]
			})

	frappe.db.commit()

	return {
		"success": len(created_items) > 0,
		"total_rows": len(rows_data),
		"created_count": len(created_items),
		"error_count": len(errors),
		"created_items": created_items,
		"errors": errors
	}


@frappe.whitelist()
def get_dashboard_metrics(calendar=None, year=None, month=None, format=None, domain=None, risk=None):
	"""Aggregates real-time metrics, KPI numbers, and chart distributions for the Executive Marketing Dashboard."""
	user = frappe.session.user
	roles = frappe.get_roles(user)
	is_lead = user == "Administrator" or "Marketing Lead" in roles or "System Manager" in roles

	if not is_lead:
		frappe.throw(_("Access Restricted: Only Marketing Leads and System Managers can access the Marketing Operations Dashboard."), frappe.PermissionError)

	conditions = []
	params = {}


	# Calendar filter
	if calendar and calendar != "All":
		conditions.append("content_calendar = %(calendar)s")
		params["calendar"] = calendar

	# Year and Month filters on planned_publish_date
	if year and str(year) != "All":
		conditions.append("YEAR(planned_publish_date) = %(year)s")
		params["year"] = int(year)

	if month and str(month) not in ["All", "0", ""]:
		try:
			m_int = int(month)
			conditions.append("MONTH(planned_publish_date) = %(month)s")
			params["month"] = m_int
		except Exception:
			pass

	# Format filter
	if format and format != "All":
		conditions.append("content_type = %(format)s")
		params["format"] = format

	# Industry Domain filter
	if domain and domain != "All":
		conditions.append("industry_domain = %(domain)s")
		params["domain"] = domain

	# Risk Status filter
	if risk and risk != "All":
		conditions.append("risk_flag = %(risk)s")
		params["risk"] = risk

	where_str = ("WHERE " + " AND ".join(conditions)) if conditions else ""

	# Fetch raw item data matching filter
	items = frappe.db.sql(f"""
		SELECT
			name, title, status, workflow_state, content_type,
			industry_domain, planned_publish_date, due_date,
			risk_flag, ai_score, assigned_to, reviewer_technical
		FROM `tabContent Item`
		{where_str}
	""", params, as_dict=True)

	# KPI Counters
	counts = {
		"total": len(items),
		"planned": 0,
		"briefed": 0,
		"in_progress": 0,
		"in_review": 0,
		"in_revision": 0,
		"approved": 0,
		"published": 0,
		"late_risk": 0,
		"avg_ai_score": 0.0
	}

	status_dist = {
		"Planned": 0, "Briefed": 0, "In Progress": 0,
		"In Review": 0, "In Revision": 0, "Approved": 0, "Published": 0
	}
	format_dist = {}
	domain_dist = {}
	risk_dist = {"On track": 0, "At risk": 0, "Late": 0}
	monthly_trend = {m: 0 for m in range(1, 13)}

	ai_scores = []

	for it in items:
		st = it.workflow_state or it.status or "Planned"
		if st in status_dist:
			status_dist[st] += 1
		else:
			status_dist[st] = 1

		# Key count mapping
		key = st.lower().replace(" ", "_")
		if key in counts:
			counts[key] += 1

		# Formats
		fmt = it.content_type or "Unassigned"
		format_dist[fmt] = format_dist.get(fmt, 0) + 1

		# Domains
		dom = it.industry_domain or "General"
		domain_dist[dom] = domain_dist.get(dom, 0) + 1

		# Risk
		rf = it.risk_flag or "On track"
		risk_dist[rf] = risk_dist.get(rf, 0) + 1
		if rf == "Late":
			counts["late_risk"] += 1

		# AI Scores
		if it.ai_score is not None and it.ai_score > 0:
			ai_scores.append(float(it.ai_score))

		# Monthly trend
		if it.planned_publish_date:
			try:
				d = getdate(it.planned_publish_date)
				monthly_trend[d.month] += 1
			except Exception:
				pass

	if ai_scores:
		counts["avg_ai_score"] = round(sum(ai_scores) / len(ai_scores), 1)

	# Active options for filter dropdowns
	calendars = frappe.get_all("Content Calendar", fields=["name", "status", "from_date", "to_date"], order_by="creation desc")
	formats = frappe.get_all("Content Item Option", filters={"option_type": "Format", "is_active": 1}, pluck="option_label")
	domains = frappe.get_all("Content Item Option", filters={"option_type": "Industry Domain", "is_active": 1}, pluck="option_label")

	month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
	monthly_series = [
		{"month": month_names[m - 1], "month_num": m, "count": monthly_trend[m]}
		for m in range(1, 13)
	]

	return {
		"kpis": counts,
		"status_distribution": status_dist,
		"format_distribution": format_dist,
		"domain_distribution": domain_dist,
		"risk_distribution": risk_dist,
		"monthly_trend": monthly_series,
		"calendars": calendars,
		"formats": formats,
		"domains": domains,
		"is_lead": is_lead
	}



@frappe.whitelist()
def trigger_ai_copilot(docname):
	"""Manual API trigger for Marketing Copilot Review (Writer) via Background Job Enqueueing."""
	if not frappe.db.exists("Content Item", docname):
		frappe.throw(_("Content Item {0} not found.").format(docname), frappe.DoesNotExistError)

	doc = frappe.get_doc("Content Item", docname)
	user = (frappe.session.user or "").lower()
	assigned = (doc.assigned_to or "").lower()
	roles = frappe.get_roles(user)
	is_lead = user == "administrator" or "Marketing Lead" in roles or "System Manager" in roles

	# Strict document-level role check: Writer copilot can only be run by the assigned writer or a lead
	if user != assigned and not is_lead:
		frappe.throw(_("Only the assigned writer ({0}) or a Marketing Lead can run Writer Copilot Review.").format(doc.assigned_to), frappe.PermissionError)

	settings = frappe.get_single("Marketing Settings")
	if not getattr(settings, "enable_ai_copilot", 0):
		frappe.throw(
			_("AI Copilot is currently disabled in Marketing Settings. Enable it before triggering a review."),
			frappe.ValidationError
		)

	# Check primary content draft is attached
	if not (doc.content_file_1 or "").strip():
		frappe.throw(
			_("<b>Primary Content Draft</b> is mandatory to run AI Copilot Review. Please attach a draft file first."),
			frappe.ValidationError
		)

	# Check usage limit for Writer
	max_reviews = int(getattr(settings, "max_writer_copilot_reviews_per_item", 2) or 2)
	current_count = len([r for r in (doc.ai_reviews or []) if r.get("review_type") == "Writer"])
	if current_count >= max_reviews:
		frappe.throw(
			_("Writer Copilot review limit reached ({0}/{1}). No additional AI reviews can be triggered for this item.").format(current_count, max_reviews),
			frappe.ValidationError
		)

	print(f"\n================================================================================", flush=True)
	print(f"🎯 [COPILOT API TRIGGERED] Content Item: {docname} | Initiated by: {user} (Writer Review)", flush=True)
	print(f"================================================================================\n", flush=True)

	from oda_marketing.oda_marketing.ai_engine.runner import run_ai_review
	try:
		run_ai_review(docname=docname, review_type="Writer")
	except Exception as e:
		frappe.log_error(f"AI Copilot execution error: {str(e)}")
		print(f"\n❌ [AI REVIEW EXECUTION ERROR] {str(e)}\n", flush=True)
		raise

	return frappe.get_doc("Content Item", docname)


@frappe.whitelist()
def trigger_reviewer_copilot(docname, instructions=None):
	"""Reviewer-triggered Copilot review with custom instructions via Immediate Execution."""
	if not frappe.db.exists("Content Item", docname):
		frappe.throw(_("Content Item {0} not found.").format(docname), frappe.DoesNotExistError)

	doc = frappe.get_doc("Content Item", docname)
	user = (frappe.session.user or "").lower()
	assigned = (doc.assigned_to or "").lower()
	reviewer = (doc.reviewer_technical or "").lower()
	roles = frappe.get_roles(user)
	is_doc_writer = (user == assigned)
	is_lead = user == "administrator" or "Marketing Lead" in roles or "System Manager" in roles

	# The writer on this document must NEVER run the Reviewer Copilot, regardless of global roles
	if is_doc_writer:
		frappe.throw(
			_("As the assigned writer, you cannot run the Reviewer Copilot on your own deliverable. Only the designated Reviewer ({0}) or a non-author Lead can do this.").format(doc.reviewer_technical or "Reviewer"),
			frappe.PermissionError
		)

	# Strict document-level role check: Reviewer copilot can only be run by the assigned reviewer or a lead (who is not the writer)
	if user != reviewer and not is_lead:
		frappe.throw(_("Only the designated Reviewer ({0}) can run Reviewer Copilot Review.").format(doc.reviewer_technical), frappe.PermissionError)

	settings = frappe.get_single("Marketing Settings")
	if not getattr(settings, "enable_ai_copilot", 0):
		frappe.throw(
			_("AI Copilot is currently disabled in Marketing Settings. Enable it before triggering a review."),
			frappe.ValidationError
		)

	# Check primary content draft is attached
	if not (doc.content_file_1 or "").strip():
		frappe.throw(
			_("<b>Primary Content Draft</b> is mandatory to run AI Copilot Review. Please attach a draft file first."),
			frappe.ValidationError
		)

	# Check usage limit for Reviewer
	max_reviews = int(getattr(settings, "max_reviewer_copilot_reviews_per_item", 2) or 2)
	current_count = len([r for r in (doc.ai_reviews or []) if r.get("review_type") == "Reviewer"])
	if current_count >= max_reviews:
		frappe.throw(
			_("Reviewer Copilot review limit reached ({0}/{1}). No additional AI reviews can be triggered for this item.").format(current_count, max_reviews),
			frappe.ValidationError
		)

	# Store reviewer instructions on the document
	if instructions:
		doc.db_set("reviewer_copilot_instructions", instructions, update_modified=False)

	print(f"\n================================================================================", flush=True)
	print(f"🎯 [COPILOT API TRIGGERED] Content Item: {docname} | Initiated by: {user} (Reviewer Review)", flush=True)
	if instructions:
		print(f"📝 Reviewer Instructions: {instructions}", flush=True)
	print(f"================================================================================\n", flush=True)

	from oda_marketing.oda_marketing.ai_engine.runner import run_ai_review
	try:
		run_ai_review(docname=docname, reviewer_instructions=instructions, review_type="Reviewer")
	except Exception as e:
		frappe.log_error(f"Reviewer Copilot execution error: {str(e)}")
		print(f"\n❌ [REVIEWER AI EXECUTION ERROR] {str(e)}\n", flush=True)
		raise

	return frappe.get_doc("Content Item", docname)




@frappe.whitelist()
def get_ai_copilot_status():
	"""Whitelisted helper to check AI Copilot status without requiring read permission on Marketing Settings DocType."""
	try:
		settings = frappe.get_single("Marketing Settings")
		return {
			"enable_ai_copilot": int(getattr(settings, "enable_ai_copilot", 0) or 0),
			"max_writer_copilot_reviews_per_item": int(getattr(settings, "max_writer_copilot_reviews_per_item", 2) or 2),
			"max_reviewer_copilot_reviews_per_item": int(getattr(settings, "max_reviewer_copilot_reviews_per_item", 2) or 2)
		}
	except Exception:
		return {"enable_ai_copilot": 0, "max_writer_copilot_reviews_per_item": 2, "max_reviewer_copilot_reviews_per_item": 2}

